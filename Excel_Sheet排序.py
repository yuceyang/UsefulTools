import os

from openpyxl import load_workbook


file_path = r"C:\Users\C-Young\Desktop\【02 系统构成】_渭南澄城光伏项目等保测评项目_电力监控系统_虞策杨（2024-06-04 094013538）\merged_excel.xlsx"
def Sorting(file_path):

    # 加载 Excel 文件
    workbook = load_workbook(file_path)

    # 定义目标工作表顺序
    target_order = ['机房',
                    '网络设备',
                    '安全设备',
                    '业务应用软件&平台',
                    '系统管理平台',
                    '服务器&存储设备',
                    '数据库管理系统',
                    '终端&感知设备&现场设备',
                    '其他系统或设备',
                    '密码产品',
                    '关键数据类别',
                    '安全相关人员',
                    '安全管理文档']

    # 获取所有工作表名称
    sheet_names = workbook.sheetnames

    # 创建一个新的工作表列表
    new_sheets = []

    # 根据目标顺序添加工作表
    for sheet_name in target_order:
        if sheet_name in sheet_names:
            new_sheets.append(workbook[sheet_name])
            sheet_names.remove(sheet_name)

    # 添加剩余的工作表
    new_sheets.extend([workbook[sheet_name] for sheet_name in sheet_names])

    # 清空原有的工作表
    workbook._sheets.clear()

    # 将新的工作表列表添加到工作簿中
    workbook._sheets.extend(new_sheets)

    # 保存修改后的 Excel 文件
    workbook.save(file_path)

# 检查文件是否存在
if os.path.exists(file_path):
    # 调用函数查找文件名
    Sorting(file_path)
else:
    print("文件不存在")