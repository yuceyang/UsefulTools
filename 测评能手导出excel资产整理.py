import os
import pandas as pd
import openpyxl
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from tkinter import Tk
from tkinter.filedialog import askdirectory
##
#   环境需要使用xlsxwriter模块
#   pip install xlsxwriter
#
# #

def start():
    # 创建一个Tkinter根窗口，并隐藏它
    root = Tk()
    root.withdraw()
    # 使用文件资源管理器获取文件夹路径
    folder_path = askdirectory(title="请选择要处理的文件夹")
    # 检查是否选择了文件夹路径
    if folder_path:
        print(f"你选择的文件夹路径是: {folder_path}")

        # 检查merged_excel.xlsx文件是否已经存在
        merged_file_path = os.path.join(folder_path, "merged_excel.xlsx")
        if os.path.exists(merged_file_path):
            print(f"文件【merged_excel.xlsx】已经存在，跳过文件处理步骤，请删除【merged_excel.xlsx】文件后重试!!!。")
        else:
            perform_replace_task(folder_path)
    else:
        print("你没有选择任何文件夹。")


def perform_replace_task(folder_path):
    #删除和替换文件名
    _find_refilename(folder_path)
    #合并excel
    _excel_hb(folder_path)
    #操作excel，删除行、列、表格美化等
    _excel_operate(folder_path+"/merged_excel.xlsx")
    #sheet排序
    _Sorting(folder_path+"/merged_excel.xlsx")

def _find_refilename(folder_path):#删除和替换文件名
    # 定义模糊查询条件和替换条件的字典
    conditions = {
        "机房": "机房",
        "网络设备": "网络设备",
        "安全设备": "安全设备",
        "业务应用软件": "业务应用软件&平台",
        "系统管理平台": "系统管理平台",
        "服务器": "服务器&存储设备",
        "终端": "终端&感知设备&现场设备",
        "其他系统或设备": "其他系统或设备",
        "数据库管理系统": "数据库管理系统",
        "关键数据类别": "关键数据类别",
        "安全相关人员": "安全相关人员",
        "密码产品": "密码产品",
        "安全管理文档": "安全管理文档"
    }
    delet_filelist = {"区域边界", "安全管理中心", "全局对象"}
    # 遍历文件夹及其子文件夹
    for root, dirs, files in os.walk(folder_path):
        for filename in files:
            try:
                # 替换流
                new_filename = None
                for query, replacement in conditions.items():
                    if query in filename:
                        new_filename = replacement + ".xlsx"
                        break
                if new_filename:
                    # 拼接完整的源文件路径和目标文件路径
                    src = os.path.join(folder_path, filename)
                    dst = os.path.join(folder_path, new_filename)
                    # 重命名文件
                    os.rename(src, dst)
                # 删除流
                for de_fi in delet_filelist:
                    if de_fi in filename:
                        os.remove(os.path.join(folder_path, filename))
                        break
            except UnicodeEncodeError:
                print(filename.encode('utf-8').decode('utf-8'))


def _excel_hb(folder_path):#合并excel
    # 获取文件夹下所有Excel文件的文件名
    excel_files = [f for f in os.listdir(folder_path) if f.endswith(".xlsx")]

    # 创建一个新的Excel写入器
    writer = pd.ExcelWriter(folder_path+"/merged_excel.xlsx", engine="xlsxwriter")

    # 遍历每个Excel文件并将其写入新的Excel文件中
    for file in excel_files:
        file_path = os.path.join(folder_path, file)
        sheet_name = os.path.splitext(file)[0]  # 使用文件名作为sheet名称

        # 读取Excel文件
        df = pd.read_excel(file_path)

        # 将数据写入新的Excel文件中
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    # 保存并关闭写入器
    writer.close()


def _excel_operate(folder_path):#操作excel，删除行、列、表格美化等
    # 打开Excel文件
    workbook = openpyxl.load_workbook(folder_path)
    # 定义每个sheet需要删除的行和列,并设置行高、列宽
    #例如：
    # '06服务器一存储设备': {'rows': [1],'cols': [8,11,12,13],'row_heights': {1: 14.4, 2: 9, 3: 16},'col_widths': {'A': 20, 'B': 15, 'C': 30}}
    sheet_deletions = {
        '机房': {'rows': [1],'cols': [5,6],'row_heights': {None: 14.4},'col_widths': {'A': 9, 'B': 19, 'C': 60 ,'D':18}},
        '网络设备': {'rows': [1],'cols': [7,10,11,12],'row_heights': {None: 14.4},'col_widths': {'A': 9, 'B': 36, 'C': 18 ,'D':35 ,'E':35 ,'F':35 ,'G':9 ,'H':18 }},
        '安全设备': {'rows': [1],'cols': [7,10,11,12],'row_heights': {None: 14.4},'col_widths':  {'A': 9, 'B': 36, 'C': 18 ,'D':35 ,'E':35 ,'F':35 ,'G':9 ,'H':18 }},
        '业务应用软件&平台': {'rows': [1],'cols': [7,8,9],'row_heights': {None: 14.4},'col_widths':  {'A': 9, 'B': 36, 'C': 85 ,'D':35 ,'E':35 ,'F':18 }},
        '系统管理平台': {'rows': [1],'cols': [6,8,9,10],'row_heights': {None: 14.4},'col_widths':  {'A': 9, 'B': 36, 'C': 18 ,'D':35 ,'E':35 ,'F':18 }},
        '服务器&存储设备': {'rows': [1],'cols': [8,11,12,13],'row_heights': {None: 14.4},'col_widths':  {'A': 9, 'B': 36, 'C': 18 ,'D':35 ,'E':35 ,'F':35,'G':22 ,'H':9,'I':18}},
        '终端&感知设备&现场设备': {'rows': [1],'cols': [9,10,11],'row_heights': {None: 14.4},'col_widths': {'A': 9, 'B': 36, 'C': 18 ,'D':35 ,'E':35 ,'F':35,'G':22 ,'H':9}},
        '数据库管理系统': {'rows': [1],'cols': [11,12,13],'row_heights': {None: 14.4},'col_widths': {'A': 9, 'B': 36, 'C': 18 ,'D':35 ,'E':35 ,'F':35,'G':35 ,'H':35,'I':18,'J':18}},
        '关键数据类别': {'rows': [1],'cols': [6,7,8,9,10,11,13,14],'row_heights': {None: 14.4},'col_widths': {'A': 9, 'B': 18, 'C': 35 ,'D':35 ,'E':35 ,'F':18}},
        '密码产品': {'rows': [1],'cols': [8],'row_heights': {None: 14.4},'col_widths': {'A': 9, 'B': 36, 'C': 36 ,'D':35 ,'E':35 ,'F':35,'G':9}},
        '安全相关人员': {'rows': [1],'cols': [6],'row_heights': {None: 14.4},'col_widths': {'A': 9, 'B': 18, 'C': 62 ,'D':23 ,'E':55 }},
        '安全管理文档': {'rows': [1],'cols': [4],'row_heights': {None: 14.4},'col_widths':{'A': 9, 'B': 50, 'C': 162 }},
        '其他系统或设备': {'rows': [1],'cols': [6,8,9,10],'row_heights': {None: 14.4},'col_widths':{'A': 9, 'B': 35, 'C': 9 ,'D':35 ,'E':35,'F':18 }}
    }

    # 遍历每个sheet并进行删除和美化操作
    for sheet_name, deletions in sheet_deletions.items():
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            # 删除指定行
            for row_idx in sorted(deletions['rows'], reverse=True):
                sheet.delete_rows(row_idx)
            # 删除指定列
            for col_idx in sorted(deletions['cols'], reverse=True):
                sheet.delete_cols(col_idx)
            # 设置全表白色填充
            for row in sheet.iter_rows():
                for cell in row:
                    cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            # 设置指定列宽
            for col_letter, width in deletions['col_widths'].items():
                sheet.column_dimensions[col_letter].width = width
            # 设置默认行高
            default_height = deletions['row_heights'].get(None)
            if default_height:
                sheet.default_row_height = default_height
            # 设置特定行高
            for row_idx, height in deletions['row_heights'].items():
                if row_idx is not None:
                    sheet.row_dimensions[row_idx].height = height
            # 设置标题加粗蓝底
            header_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font

            # 设置单元格内外边框
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            for row in sheet.iter_rows():
                for cell in row:
                    cell.border = border

            # 设置单元格字体居中
            for row in sheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        # ... 其他处理逻辑保持不变
        else:
            print(f"警告：Sheet '{sheet_name}' 工作簿中不存在，跳过此工作表的处理。")
    # 保存修改后的Excel文件
    workbook.save(folder_path)


def _Sorting(file_path):#sheet排序

    # 加载 Excel 文件
    workbook = load_workbook(file_path)

    # 定义目标工作表顺序
    target_order = ['机房',
                    '网络设备',
                    '安全设备',
                    '业务应用软件&平台',
                    '系统管理平台',
                    '服务器&存储设备',
                    '数据库管理系统',
                    '终端&感知设备&现场设备',
                    '其他系统或设备',
                    '密码产品',
                    '关键数据类别',
                    '安全相关人员',
                    '安全管理文档']

    # 获取所有工作表名称
    sheet_names = workbook.sheetnames

    # 创建一个新的工作表列表
    new_sheets = []

    # 根据目标顺序添加工作表
    for sheet_name in target_order:
        if sheet_name in sheet_names:
            new_sheets.append(workbook[sheet_name])
            sheet_names.remove(sheet_name)

    # 添加剩余的工作表
    new_sheets.extend([workbook[sheet_name] for sheet_name in sheet_names])

    # 清空原有的工作表
    workbook._sheets.clear()

    # 将新的工作表列表添加到工作簿中
    workbook._sheets.extend(new_sheets)

    # 保存修改后的 Excel 文件
    workbook.save(file_path)

# 主程序循环
while True:
    start()

    # 提示用户选择下一步操作
    choice = input("\n按 1 重新进行下一次任务,按其他任意键输入则退出程序: ")

    if choice != "1":
        break
print("程序已退出。")



##
# 打包exe执行，注意pip install依赖包
# pyinstaller --onefile --hidden-import pandas --hidden-import openpyxl --hidden-import openpyxl.reader.excel --hidden-import openpyxl.styles --hidden-import openpyxl.cell --hidden-import tkinter --hidden-import tkinter.filedialog 测评能手导出excel资产整理.py
# #