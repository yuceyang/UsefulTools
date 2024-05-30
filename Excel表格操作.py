import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

folder_path = r"C:\Users\C-Young\Desktop\【02 系统构成】_2023西安市出租汽车管理处_西安市出租汽车服务管理信息系统_虞策杨（2024-05-30 141618801）\merged_excel.xlsx"
# 打开Excel文件
workbook = openpyxl.load_workbook(folder_path)
# 定义每个sheet需要删除的行和列,并设置行高、列宽
#例如：
# '06服务器一存储设备': {'rows': [1],'cols': [8,11,12,13],'row_heights': {1: 14.4, 2: 9, 3: 16},'col_widths': {'A': 20, 'B': 15, 'C': 30}}
sheet_deletions = {
    '01机房': {'rows': [1],'cols': [5,6],'row_heights': {None: 14.4},'col_widths': {'A': 9, 'B': 19, 'C': 60 ,'D':18}},
    '02网络设备': {'rows': [1],'cols': [7,10,11,12],'row_heights': {None: 14.4},'col_widths': {'A': 9, 'B': 36, 'C': 18 ,'D':35 ,'E':35 ,'F':35 ,'G':9 ,'H':18 }},
    '03安全设备': {'rows': [1],'cols': [7,10,11,12],'row_heights': {None: 14.4},'col_widths':  {'A': 9, 'B': 36, 'C': 18 ,'D':35 ,'E':35 ,'F':35 ,'G':9 ,'H':18 }},
    '04业务应用软件一平台': {'rows': [1],'cols': [7,8,9],'row_heights': {None: 14.4},'col_widths':  {'A': 9, 'B': 36, 'C': 18 ,'D':35 ,'E':35 ,'F':18 }},
    '05系统管理平台一全局扩展': {'rows': [1],'cols': [6,8,9,10],'row_heights': {None: 14.4},'col_widths':  {'A': 9, 'B': 36, 'C': 18 ,'D':35 ,'E':35 ,'F':18 }},
    '06服务器一存储设备': {'rows': [1],'cols': [8,11,12,13],'row_heights': {None: 14.4},'col_widths':  {'A': 9, 'B': 36, 'C': 18 ,'D':35 ,'E':35 ,'F':35,'G':22 ,'H':9,'I':18}},
    '07终端一感知设备一现场设备': {'rows': [1],'cols': [9,10,11],'row_heights': {None: 14.4},'col_widths': {'A': 9, 'B': 36, 'C': 18 ,'D':35 ,'E':35 ,'F':35,'G':22 ,'H':9}},
    '08数据库管理系统': {'rows': [1],'cols': [11,12,13],'row_heights': {None: 14.4},'col_widths': {'A': 9, 'B': 36, 'C': 18 ,'D':35 ,'E':35 ,'F':35,'G':35 ,'H':35,'I':18,'J':18}},
    '09关键数据类别': {'rows': [1],'cols': [6,7,8,9,10,11,13,14],'row_heights': {None: 14.4},'col_widths': {'A': 9, 'B': 18, 'C': 35 ,'D':35 ,'E':35 ,'F':18}},
    '10密码产品': {'rows': [1],'cols': [8],'row_heights': {None: 14.4},'col_widths': {'A': 9, 'B': 36, 'C': 36 ,'D':35 ,'E':35 ,'F':35,'G':9}},
    '11安全相关人员': {'rows': [1],'cols': [6],'row_heights': {None: 14.4},'col_widths': {'A': 9, 'B': 18, 'C': 62 ,'D':23 ,'E':55 }},
    '12安全管理文档': {'rows': [1],'cols': [4],'row_heights': {None: 14.4},'col_widths':{'A': 9, 'B': 50, 'C': 162 }}
}

# 遍历每个sheet并进行删除和美化操作
for sheet_name, deletions in sheet_deletions.items():
    print(sheet_name)
    sheet = workbook[sheet_name]
    # 删除指定行
    for row_idx in sorted(deletions['rows'], reverse=True):
        sheet.delete_rows(row_idx)
    # 删除指定列
    for col_idx in sorted(deletions['cols'], reverse=True):
        sheet.delete_cols(col_idx)
    # 设置全表白色填充
    for row in sheet.iter_rows():
        for cell in row:
            cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    # 设置指定列宽
    for col_letter, width in deletions['col_widths'].items():
        sheet.column_dimensions[col_letter].width = width
    # 设置默认行高
    default_height = deletions['row_heights'].get(None)
    if default_height:
        sheet.default_row_height = default_height
    # 设置特定行高
    for row_idx, height in deletions['row_heights'].items():
        if row_idx is not None:
            sheet.row_dimensions[row_idx].height = height
    # 设置标题加粗蓝底
    header_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    # 设置单元格内外边框
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = border

    # 设置单元格字体居中
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

# 保存修改后的Excel文件
workbook.save(folder_path)